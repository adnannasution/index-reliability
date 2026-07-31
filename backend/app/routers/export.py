"""Ekspor hasil ke Excel (multi-sheet) dan CSV per tabel (§9: tombol ekspor)."""

from __future__ import annotations

import csv
import io
import random
from datetime import date, timedelta
from typing import Any, Dict, List

import openpyxl
import pandas as pd
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from starlette.concurrency import run_in_threadpool

from app.core import store
from app.core.criticality import assess_all, build_module_j
from app.routers.analysis import get_analysis, resolve_config
from app.schemas.models import AnalyzeRequest

router = APIRouter(prefix="/api/export", tags=["ekspor"])

#: Nama sheet Excel maksimal 31 karakter.
_MAX_SHEET = 31


def _sheets_from_analysis(res, cfg) -> Dict[str, List[Dict[str, Any]]]:
    """Susun tabel-tabel hasil menjadi kumpulan sheet, mengikuti penamaan MATLAB."""
    sheets: Dict[str, List[Dict[str, Any]]] = {
        "A_KualitasData": res.data_quality,
        "B_Equipment": res.equipment,
        "B2_FunctionalSystem": res.functional_system,
        "B4_Pareto": res.pareto.get("equipment", []),
        "C_WeibullKelas": [
            {k: v for k, v in r.items() if not k.startswith("plot_")}
            for r in res.weibull_class
        ],
        "C2_WeibullFS": [
            {k: v for k, v in r.items() if not k.startswith("plot_")}
            for r in res.weibull_fs
        ],
        "C3_BetaEquipment": res.weibull_equipment,
        "D_TrenLaplace": res.laplace,
        "E1_StageRBD": [
            {k: (", ".join(v) if isinstance(v, list) else v) for k, v in r.items()}
            for r in res.rbd_stages
        ],
        "E2_SistemRBD": res.rbd_system,
        "E3_KeandalanFS": res.reliability_horizons,
        "E4_UmurKeandalan": res.reliability_life,
        "E5_Ketersediaan": res.availability,
    }
    # Tabulasi silang FS x kelas menjadi bentuk tabel.
    ct = res.crosstab
    if ct:
        rows = []
        for i, rname in enumerate(ct.get("rows", [])):
            row: Dict[str, Any] = {"Sistem": rname}
            for j, cname in enumerate(ct.get("classes", [])):
                row[cname] = ct["values"][i][j]
            rows.append(row)
        sheets["B3_TabSilang"] = rows
    if res.mrr_example:
        sheets["C3_ContohMRR"] = res.mrr_example.get("rows", [])
    return sheets


@router.post("/excel")
async def export_excel(req: AnalyzeRequest) -> StreamingResponse:
    """Unduh seluruh tabel hasil sebagai satu workbook Excel."""
    cfg = resolve_config(req.config)

    def work() -> bytes:
        res = get_analysis(req.dataset_id, cfg)
        sheets = _sheets_from_analysis(res, cfg)

        # Modul J bila dapat dihitung.
        try:
            ds = store.get(req.dataset_id)
            cost_by_tag = ds.cost_model.by_tag if ds and ds.cost_model else None
            ass = assess_all(
                res.eq_params, res.tags, res.n_cm, res.kelas, res.crit, res.fs_of_tag,
                res.beta_crow_by_tag, res.p_class_laplace, cfg, cost_by_tag=cost_by_tag,
            )
            j = build_module_j(ass, cfg)
            sheets["J1_Kekritisan"] = j["kekritisan"]
            sheets["J2_Strategi"] = j["strategi"]
            sheets["J3_SukuCadang"] = j["suku_cadang"]
            sheets["J4_DaftarTindakan"] = [
                {k: v for k, v in r.items() if k != "narasi"} for r in j["tindakan"]
            ]
        except Exception:  # noqa: BLE001 - ekspor tidak boleh gagal karena satu modul
            pass

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="xlsxwriter") as xl:
            book = xl.book
            head = book.add_format(
                {"bold": True, "bg_color": "#1F3864", "font_color": "white",
                 "border": 1, "align": "center", "valign": "vcenter", "text_wrap": True}
            )
            # Sampul.
            info = pd.DataFrame(
                [
                    ["Laporan", "Analisis Keandalan & Optimasi Pemeliharaan CDU"],
                    ["Unit", f"RU VI Balongan: Unit {cfg.unit.rstrip('-')} (CDU)"],
                    ["Jendela pengamatan", f"{res.meta['t_start']} s.d. {res.meta['t_end']}"],
                    ["T_obs (jam)", res.meta["t_obs_hours"]],
                    ["T_obs (hari)", round(res.meta["t_obs_days"] or 0, 1)],
                    ["Mode akhir jendela", res.meta["window_end"]],
                    ["Sumber beta RBD", cfg.beta_source],
                    ["Jumlah tag", res.meta["n_tags"]],
                    ["Kegagalan korektif", res.meta["total_cm"]],
                    ["Dibuat oleh", "Reliability & Maintenance Optimization Dashboard"],
                ],
                columns=["Keterangan", "Nilai"],
            )
            info.to_excel(xl, sheet_name="Sampul", index=False)
            for c, name in enumerate(info.columns):
                xl.sheets["Sampul"].write(0, c, name, head)
            xl.sheets["Sampul"].set_column(0, 0, 26)
            xl.sheets["Sampul"].set_column(1, 1, 58)

            for name, rows in sheets.items():
                if not rows:
                    continue
                df = pd.DataFrame(rows)
                sheet = name[:_MAX_SHEET]
                df.to_excel(xl, sheet_name=sheet, index=False)
                ws = xl.sheets[sheet]
                for c, col in enumerate(df.columns):
                    ws.write(0, c, str(col), head)
                    width = max(11, min(42, int(df[col].astype(str).str.len().max() or 11) + 2))
                    ws.set_column(c, c, width)
                ws.freeze_panes(1, 0)
                ws.autofilter(0, 0, max(len(df), 1), max(len(df.columns) - 1, 0))
        return buf.getvalue()

    data = await run_in_threadpool(work)
    stamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Hasil_Keandalan_CDU_{stamp}.xlsx"'},
    )


@router.get("/template")
async def download_template() -> StreamingResponse:
    """Unduh template Excel berformat SAP yang siap diisi data notifikasi dan order."""

    def build() -> bytes:
        rng = random.Random(42)
        tags = ["11-CDU-E-001", "11-CDU-E-002", "11-CDU-P-001",
                "11-CDU-P-002", "11-CDU-V-001"]
        equip = ["E-101A", "E-101B", "P-101A", "P-101B", "V-101"]
        types = ["M1", "M1", "M2", "M3", "M3"]
        crits = ["H", "H", "M", "L", "M"]
        descs = ["Kebocoran paking pompa", "Getaran tinggi kompresor",
                 "Valve bocor", "Inspeksi rutin", "Preventive maintenance"]

        def ds(d: date) -> str:
            return d.strftime("%d.%m.%Y")

        BN, BO = 2100000000, 3100000000

        notif_rows = []
        d = date(2022, 1, 10)
        for i in range(20):
            ti = i % len(tags)
            t = types[i % len(types)]
            notif_rows.append({
                "Notification": BN + i,
                "Notifictn type": t,
                "Notif.date": ds(d),
                "Functional Loc.": tags[ti],
                "Equipment": equip[ti],
                "Criticality": crits[ti],
                "Description": descs[i % len(descs)],
                "Order": BO + i if t in ("M1", "M2") else "",
            })
            d += timedelta(days=rng.randint(10, 40))

        order_rows = []
        for i in range(15):
            ti = i % len(tags)
            planned = round(rng.uniform(5e6, 120e6), -3)
            actual = round(planned * rng.uniform(0.8, 1.25), -3)
            bs = date(2022, 1, 15) + timedelta(days=i * 14)
            bf = bs + timedelta(days=rng.randint(2, 10))
            af = bf + timedelta(days=rng.randint(0, 4))
            order_rows.append({
                "Order": BO + i,
                "Order Type": "PM01" if types[i % len(types)] == "M3" else "PM02",
                "Notification": BN + i if types[i % len(types)] in ("M1", "M2") else "",
                "Functional Loc.": tags[ti],
                "Equipment": equip[ti],
                "Description": descs[i % len(descs)],
                "MaintActivType": "PM" if i % 3 == 0 else "CM",
                "Priority": rng.choice(["1", "2", "3"]),
                "PriorityType": "PM71",
                "Bas. start date": ds(bs),
                "Basic fin. date": ds(bf),
                "Actual Finish": ds(af),
                "TotalPlnndCosts": planned,
                "Total act.costs": actual,
            })

        wb = openpyxl.Workbook()

        def _hdr(ws, fill: str) -> None:
            fill_obj = PatternFill("solid", fgColor=fill)
            for cell in ws[1]:
                cell.fill = fill_obj
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[1].height = 26

        def _data_style(ws) -> None:
            alt = PatternFill("solid", fgColor="D9E1F2")
            for i, row in enumerate(ws.iter_rows(min_row=2)):
                for cell in row:
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(vertical="center")
                    if i % 2 == 1:
                        cell.fill = alt

        def _auto_w(ws) -> None:
            for col in ws.columns:
                w = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, 10), 36)

        # Sheet 1 – Notifikasi
        ws_n = wb.active
        ws_n.title = "Notifikasi"
        hdr_n = list(notif_rows[0].keys())
        ws_n.append(hdr_n)
        for r in notif_rows:
            ws_n.append([r[h] for h in hdr_n])
        _hdr(ws_n, "1F4E79")
        _data_style(ws_n)
        _auto_w(ws_n)
        ws_n.freeze_panes = "A2"

        # Sheet 2 – Order Biaya
        ws_o = wb.create_sheet("Order Biaya")
        hdr_o = list(order_rows[0].keys())
        ws_o.append(hdr_o)
        for r in order_rows:
            ws_o.append([r[h] for h in hdr_o])
        _hdr(ws_o, "375623")
        _data_style(ws_o)
        _auto_w(ws_o)
        ws_o.freeze_panes = "A2"

        # Sheet 3 – Petunjuk
        ws_p = wb.create_sheet("Petunjuk")
        ws_p.column_dimensions["A"].width = 22
        ws_p.column_dimensions["B"].width = 62
        ws_p.column_dimensions["C"].width = 12

        guide = [
            ("PETUNJUK PENGGUNAAN TEMPLATE", "", ""),
            ("", "", ""),
            ("SHEET: NOTIFIKASI", "", ""),
            ("Kolom", "Keterangan", "Wajib?"),
            ("Notification", "Nomor notifikasi SAP (angka)", "Opsional"),
            ("Notifictn type", "M1 / M2 = kegagalan korektif · M3 = preventive maintenance", "WAJIB"),
            ("Notif.date", "Tanggal notifikasi. Format DD.MM.YYYY atau serial Excel", "WAJIB"),
            ("Functional Loc.", "Kode lokasi SAP, harus diawali prefix unit (default: 11-)", "WAJIB"),
            ("Equipment", "Nomor equipment SAP", "Opsional"),
            ("Criticality", "Tingkat kritikal: H / M / L", "Opsional"),
            ("Description", "Deskripsi notifikasi", "Opsional"),
            ("Order", "Nomor order terkait", "Opsional"),
            ("", "", ""),
            ("SHEET: ORDER BIAYA", "", ""),
            ("Kolom", "Keterangan", "Wajib?"),
            ("Order", "Nomor order SAP", "Opsional"),
            ("Order Type", "Tipe order (PM01, PM02, dst.)", "Opsional"),
            ("Notification", "Nomor notifikasi terkait", "Opsional"),
            ("Functional Loc.", "Kode lokasi SAP, harus diawali prefix unit", "WAJIB"),
            ("Equipment", "Nomor equipment SAP", "Opsional"),
            ("Description", "Deskripsi order", "Opsional"),
            ("MaintActivType", "PM = preventif · CM = korektif", "Opsional"),
            ("Priority", "Prioritas (1 / 2 / 3)", "Opsional"),
            ("PriorityType", "Tipe prioritas (mis. PM71)", "Opsional"),
            ("Bas. start date", "Tanggal mulai rencana", "Opsional"),
            ("Basic fin. date", "Tanggal selesai rencana", "Opsional"),
            ("Actual Finish", "Tanggal selesai aktual", "Opsional"),
            ("TotalPlnndCosts", "Total biaya terencana (Rupiah)", "WAJIB"),
            ("Total act.costs", "Total biaya aktual (Rupiah)", "Opsional"),
            ("", "", ""),
            ("CATATAN", "", ""),
            ("", "Prefix unit default: '11-' (CDU RU VI Balongan Unit 11).", ""),
            ("", "Bila prefix berbeda (mis. 12-), ubah di menu Pengaturan metode.", ""),
            ("", "Kedua sheet bisa dalam satu file; nama sheet bebas.", ""),
            ("", "Sistem mendeteksi jenis berkas dari nama kolom, bukan nama file.", ""),
        ]

        for ri, row in enumerate(guide, 1):
            for ci, val in enumerate(row, 1):
                cell = ws_p.cell(row=ri, column=ci, value=val)
                cell.font = Font(size=10)
                cell.alignment = Alignment(wrap_text=True, vertical="center")

        ws_p[1][0].font = Font(bold=True, size=13, color="1F4E79")
        for cell in ws_p[3]:
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
        for cell in ws_p[4]:
            cell.fill = PatternFill("solid", fgColor="2E75B6")
            cell.font = Font(bold=True, color="FFFFFF")
        for cell in ws_p[14]:
            cell.fill = PatternFill("solid", fgColor="375623")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
        for cell in ws_p[15]:
            cell.fill = PatternFill("solid", fgColor="548235")
            cell.font = Font(bold=True, color="FFFFFF")
        ws_p[31][0].font = Font(bold=True, color="C00000", size=10)
        red = Font(bold=True, color="C00000", size=10)
        for ri, row in enumerate(guide, 1):
            if len(row) > 2 and row[2] == "WAJIB":
                ws_p.cell(row=ri, column=3).font = red

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    data = await run_in_threadpool(build)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Template_SAP_CDU.xlsx"'},
    )


@router.post("/csv/{table}")
async def export_csv(table: str, req: AnalyzeRequest) -> StreamingResponse:
    """Unduh satu tabel sebagai CSV (pemisah titik-koma, ramah Excel Indonesia)."""
    cfg = resolve_config(req.config)

    def work() -> str:
        res = get_analysis(req.dataset_id, cfg)
        sheets = _sheets_from_analysis(res, cfg)
        key = next((k for k in sheets if k.lower() == table.lower()), None)
        if key is None:
            raise HTTPException(
                status_code=404,
                detail=f"Tabel '{table}' tidak dikenal. Pilihan: {', '.join(sheets.keys())}",
            )
        rows = sheets[key]
        if not rows:
            raise HTTPException(status_code=404, detail=f"Tabel '{table}' kosong.")
        out = io.StringIO()
        w = csv.DictWriter(out, fieldnames=list(rows[0].keys()), delimiter=";",
                           extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
        return out.getvalue()

    text = await run_in_threadpool(work)
    return StreamingResponse(
        io.StringIO(text), media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{table}.csv"'},
    )
